#Importing neccessary libraries
import requests
import pandas as pd
import openpyxl
import time
from datetime import datetime
from fpdf import FPDF
import matplotlib.pyplot as plt
import seaborn as sns
import os

# Function to fetch live cryptocurrency data
def fetch_crypto_data():
    try:
        url = "https://api.coingecko.com/api/v3/coins/markets"
        params = {
            "vs_currency": "usd",
            "order": "market_cap_desc",
            "per_page": 50,
            "page": 1,
            "sparkline": False,
        }
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        print(f"Error fetching data: {e}")
        return []

# Function to analyze the data
def analyze_data(data):
    df = pd.DataFrame(data)
    # Add timestamp column
    df['Last Updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Extracting key columns
    df = df[["name", "symbol", "current_price", "market_cap", "total_volume",
             "price_change_percentage_24h", "high_24h", "low_24h", "Last Updated"]]
    df.columns = ["Name", "Symbol", "Current Price (USD)", "Market Cap", "24h Trading Volume",
                 "24h Price Change (%)", "24h High", "24h Low", "Last Updated"]

    # Additional Analysis
    market_dominance = (df["Market Cap"] / df["Market Cap"].sum()) * 100
    df["Market Dominance (%)"] = market_dominance

    # Volume to Market Cap Ratio (Trading intensity)
    df["Volume/Market Cap Ratio"] = df["24h Trading Volume"] / df["Market Cap"]

    # Price Volatility (High-Low Range)
    df["24h Price Range (%)"] = ((df["24h High"] - df["24h Low"]) / df["24h Low"]) * 100

    # Comprehensive Analysis
    analysis = {
        "Top 5 by Market Cap": df.nlargest(5, "Market Cap"),
        "Average Price": df["Current Price (USD)"].mean(),
        "Highest 24h Change (%)": df["24h Price Change (%)"].max(),
        "Lowest 24h Change (%)": df["24h Price Change (%)"].min(),
        "Total Market Cap": df["Market Cap"].sum(),
        "Average Volume": df["24h Trading Volume"].mean(),
        "Most Volatile": df.nlargest(5, "24h Price Range (%)")[["Name", "24h Price Range (%)"]],
        "Highest Volume/Market Cap": df.nlargest(5, "Volume/Market Cap Ratio")[["Name", "Volume/Market Cap Ratio"]],
        "Market Leaders": df.nlargest(3, "Market Dominance (%)")[["Name", "Market Dominance (%)"]]
    }
    return df, analysis

# Function to generate analysis report
def generate_report(df, analysis):
    report_path = "/home/Errorsearching/static/crypto_analysis_report.pdf"
    pdf = FPDF()
    pdf.add_page()

    # Title
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Cryptocurrency Market Analysis Report', ln=True, align='C')
    pdf.set_font('Arial', 'I', 10)
    pdf.cell(0, 10, f'Generated on {datetime.now().strftime("%Y-%m-%d")}', ln=True, align='C')

    # Market Overview
    pdf.set_font('Arial', 'B', 14)
    pdf.ln(10)
    pdf.cell(0, 10, 'Market Overview', ln=True)
    pdf.set_font('Arial', '', 12)

    total_market_cap = analysis["Total Market Cap"] / 1e9  # Convert to billions
    pdf.cell(0, 10, f'Total Market Cap: ${total_market_cap:.2f}B', ln=True)
    pdf.cell(0, 10, f'Average 24h Volume: ${analysis["Average Volume"]/1e9:.2f}B', ln=True)

    # Market Leaders
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 10, 'Market Leaders', ln=True)
    pdf.set_font('Arial', '', 12)
    for _, row in analysis["Market Leaders"].iterrows():
        pdf.cell(0, 10, f'{row["Name"]}: {row["Market Dominance (%)"]:.2f}% market share', ln=True)

    # Price Performance
    pdf.ln(10)
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 10, '24h Price Performance', ln=True)
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 10, f'Highest Change: {analysis["Highest 24h Change (%)"]:.2f}%', ln=True)
    pdf.cell(0, 10, f'Lowest Change: {analysis["Lowest 24h Change (%)"]:.2f}%', ln=True)

    # Save report
    try:
        pdf.output(report_path)
        print(f"Analysis report generated at {report_path}")
    except Exception as e:
        print(f"Error generating report: {e}")

def update_excel(dataframe, analysis):
    filename = "/home/Errorsearching/static/live_crypto_data.xlsx"
    try:
        with pd.ExcelWriter(filename, engine="openpyxl", mode="w") as writer:
            dataframe.to_excel(writer, sheet_name="Live Data", index=False)

            # Enhanced summary sheet
            summary = pd.DataFrame({
                "Metric": [
                    "Average Price (USD)",
                    "Highest 24h Change (%)",
                    "Lowest 24h Change (%)",
                    "Total Market Cap (USD)",
                    "Average Trading Volume (USD)",
                ],
                "Value": [
                    analysis["Average Price"],
                    analysis["Highest 24h Change (%)"],
                    analysis["Lowest 24h Change (%)"],
                    analysis["Total Market Cap"],
                    analysis["Average Volume"]
                ]
            })
            summary.to_excel(writer, sheet_name="Analysis", index=False)

            # Market leaders sheet
            analysis["Top 5 by Market Cap"].to_excel(writer, sheet_name="Top 5 Market Cap", index=False)

            # Volatility sheet
            analysis["Most Volatile"].to_excel(writer, sheet_name="Volatility Analysis", index=False)

        print(f"Excel file updated at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    except Exception as e:
        print(f"Error updating Excel file: {e}")

def main():
    print("Starting cryptocurrency data updater...")
    while True:
        try:
            print(f"\nFetching data at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            data = fetch_crypto_data()
            if data:
                dataframe, analysis = analyze_data(data)
                update_excel(dataframe, analysis)
                generate_report(dataframe, analysis)  # Generate PDF report
            else:
                print("No data fetched. Retrying in 5 minutes...")
            time.sleep(300)
        except KeyboardInterrupt:
            print("\nStopping the updater...")
            break
        except Exception as e:
            print(f"Unexpected error: {e}")
            print("Retrying in 5 minutes...")
            time.sleep(300)

if __name__ == "__main__":
    main()
